from openpyxl import load_workbook
import mysql.connector


def catalogar(*args):
    dict_dados = {}
    lista_dados = []
    lista_icj = []
    contador = args[0].max_row
    for linha in range(2, contador + 1):
        icj = args[0].cell(row=linha, column=1).value
        lista_icj.append(icj)
        lista_coluna = [5, 7]
        for coluna in lista_coluna:
            info = args[0].cell(row=linha, column=coluna).value
            lista_dados.append(info)
        dict_dados[icj] = lista_dados
        dict_dados.copy()
        lista_dados = []
    return lista_icj, dict_dados


def extrair_matricula(*args):
    lista_matricula = []
    connection = mysql.connector.connect(host='localhost', database='bd_contratos', user='root', password='', port=3306)
    with connection.cursor() as cur:
        query = ('SELECT matricula FROM empregados WHERE chave=%s')
        for n in range(0, 2):
            cond = args[0][n]
            cur.execute(query, [cond])
            myresult = cur.fetchall()
            lista_matricula.append(myresult[0][0])
        return lista_matricula




def transferir():
    wb = load_workbook(filename='Planilha Guia_dados.xlsx')
    sheet = wb['Info Contrato']
    dados_relevantes = catalogar(sheet)
    icj = sorted(set(dados_relevantes[0]))
    for valor in icj:
        info = dados_relevantes[1][valor]
        matriculas = extrair_matricula(info)


if __name__ == '__main__':
    transferir()
