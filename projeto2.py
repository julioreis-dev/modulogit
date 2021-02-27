from openpyxl import load_workbook
import mysql.connector
import test_connection


def conectar_bd(*args):
    dict_gerais = args[1]
    connection = mysql.connector.connect(host=args[2], database=args[3], user=args[4], password=args[5], port=args[6])
    for chaves in args[0]:
        extracao = dict_gerais[chaves]
        sql = 'INSERT INTO empregados (chave, nome, matricula, função, email, telefone) ' \
              'VALUES (%s, %s, %s, %s, %s, %s)'
        values = (chaves, extracao[4], extracao[2], extracao[3], extracao[0], extracao[5])
        cursor = connection.cursor()
        cursor.execute(sql, values)
        connection.commit()
        print(cursor.rowcount, f'Registro da chave {chaves} inserido com sucesso.')
    connection.close()


def catalogar(*args):
    dict_dados = {}
    lista_dados = []
    lista_chave = []
    contador = args[0].max_row
    for linha in range(2, contador + 1):
        icj = args[0].cell(row=linha, column=2).value
        lista_chave.append(icj)
        lista_coluna = [1, 2, 7, 3, 4, 5]
        for coluna in lista_coluna:
            info = args[0].cell(row=linha, column=coluna).value
            lista_dados.append(info)
        dict_dados[icj] = lista_dados
        dict_dados.copy()
        lista_dados = []
    return lista_chave, dict_dados


def main():
    wb = load_workbook(filename='Planilha Guia_dados.xlsx')
    sheet = wb['Chaves']
    teste = test_connection.testar_bd('localhost', 'bd_contratos', 'root', '', 3306)
    if teste == 'ok':
        dados_contratual = catalogar(sheet)
        print('Teste de conexão com o banco de dados estabelecida com sucesso!')
        nova_lista = sorted(set(dados_contratual[0]))
        conectar_bd(nova_lista, dados_contratual[1], 'localhost', 'bd_contratos', 'root', '', 3306)
        wb.close()
    else:
        print('Conexão não estabelecida!')
        exit()


if __name__ == '__main__':
    main()
