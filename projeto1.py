from openpyxl import load_workbook
import mysql.connector
import test_connection


def create_table_mysql_db(*args):
    connection = mysql.connector.connect(host=args[0], database=args[1], user=args[2], password=args[3], port=args[4])
    sql = 'CREATE TABLE IF NOT EXISTS my_contratos ' \
          '(id INT AUTO_INCREMENT PRIMARY KEY,' \
          'icj VARCHAR(20) NOT NULL,' \
          'equipamento VARCHAR(10), ' \
          'embarcacao VARCHAR(50) NOT NULL,' \
          'chave_gerente VARCHAR(4),' \
          'gerente VARCHAR(20),' \
          'chave_fiscal VARCHAR(4),' \
          'fiscal VARCHAR(20),' \
          'empresa VARCHAR(100) NOT NULL,' \
          'inicio DATE,' \
          'termino DATE, ' \
          'cessão VARCHAR(10));'
    cursor = connection.cursor()
    cursor.execute(sql)
    connection.commit()
    print("Opened database successfully")
    return 'Tabela criada com sucesso'


def conectar_bd(*args):
    dict_gerais = args[1]
    connection = mysql.connector.connect(host=args[2], database=args[3], user=args[4], password=args[5], port=args[6])
    for contrato_icj in args[0]:
        extracao = dict_gerais[contrato_icj]
        sql = 'INSERT INTO my_contratos (icj, equipamento, embarcacao, chave_gerente, gerente,' \
              ' chave_fiscal, fiscal, empresa, inicio, termino, cessão) ' \
              'VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)'
        values = (contrato_icj, extracao[0], extracao[1], extracao[2], extracao[3], extracao[4],
                  extracao[5], extracao[6], extracao[7], extracao[8], extracao[9])
        cursor = connection.cursor()
        cursor.execute(sql, values)
        connection.commit()
        print(cursor.rowcount, f'Registro da embarcação {extracao[1]} foi inserido com sucesso.')
    connection.close()


def catalogar(*args):
    dict_dados = {}
    contador = args[0].max_row
    lista_icj = [args[0].cell(row=line_excel, column=1).value for line_excel in range(2, contador + 1)]
    for line in range(2, contador + 1):
        dict_dados[args[0].cell(row=line, column=1).value] = [args[0].cell(row=line, column=coluna).value
                                                              for coluna in [2, 4, 5, 6, 7, 8, 9, 10, 11, 12]]
        dict_dados.copy()
    return lista_icj, dict_dados


def delete_table(*args):
    connection = mysql.connector.connect(host=args[0], database=args[1], user=args[2], password=args[3], port=args[4])
    sql = 'DROP TABLE my_contratos;'
    cur = connection.cursor()
    cur.execute(sql)
    connection.commit()
    print('Planilha deletada com sucesso!!!')
    connection.close()


def main():
    wb = load_workbook(filename='Planilha Guia_dados.xlsx')
    sheet = wb['Info Contrato']
    teste = test_connection.testar_bd('localhost', 'bd_contratos', 'root', 'root', 3306)
    if teste:
        print('Teste de conexão com o banco de dados, realizada com sucesso!!!')
        dados_contratual = catalogar(sheet)
        lista_ordenada = sorted(set(dados_contratual[0]))
        create_table_mysql_db('localhost', 'bd_contratos', 'root', 'root', 3306)
        conectar_bd(lista_ordenada, dados_contratual[1], 'localhost', 'bd_contratos', 'root', 'root', 3306)
        wb.close()
    else:
        print('Conexão não estabelecida!')
        exit()


# delete_table('localhost', 'bd_contratos', 'root', 'root', 3306)

if __name__ == '__main__':
    main()
